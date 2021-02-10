import openpyxl

def smallest2(a,b):
    if a != 0 and b != 0:
        if a == b:
            return 1
        else:
            if a < b:
                return 1
            elif b < a:
                return 2
    else:
        return -1


def smallest(a,b,c):
    if a != 0 and b != 0 and c != 0:
        if a == b == c:
            return 1
        else:
            if a < b:
                if a < c:
                    return 1
                elif c < a:
                    return 3
                else:
                    return 1
            elif b < a:
                if b < c:
                    return 2
                elif c < b:
                    return 3
                else:
                    return 2
            elif a == b:
                if a < c:
                    return 1
                elif c < a:
                    return 3
                else:
                    return 1
            else:
                return -1
    else:
        return -1

def EXTRACTDATA():
    datasave = openpyxl.load_workbook('abc.xlsx')
    datasheet = datasave.active

    orig  = openpyxl.load_workbook(r'C:\Users\DSLR\Desktop\JAMES SIR\JMJ.xlsx')
    origsheet = orig.active

    wb1 = openpyxl.load_workbook(r'C:\Users\DSLR\Desktop\JAMES SIR\Original.xlsx')
    #print(wb.sheetnames)
      #if Q1 != ans1:
        #    print(Q1)
    #ws['C6'] = "hahaha"
    #
    #wb.save("abc.xlsx")
    ws1 = wb1.active

    wb2 = openpyxl.load_workbook(r'C:\Users\DSLR\Desktop\JAMES SIR\correct.xlsx')
    #print(wb.sheetnames)
    ws2 = wb2.active
    empty = ws1["K8"].value
    for i in range(2,199):
        cellO = "C"+str(i)
        Q1 = ws1[cellO].value
        cellO = "D" + str(i)
        Q2 = ws1[cellO].value
        cellO = "E" + str(i)
        Q3 = ws1[cellO].value
        cellO = "F" + str(i)
        Q4 = ws1[cellO].value
        cellO = "G" + str(i)
        Q5 = ws1[cellO].value
        cellO = "H" + str(i)
        Q6 = ws1[cellO].value
        cellO = "I" + str(i)
        Q7 = ws1[cellO].value
        cellO = "J" + str(i)
        Q8 = ws1[cellO].value
        cellO = "K" + str(i)
        Q9 = ws1[cellO].value
        cellO = "L" + str(i)
        Q10 = ws1[cellO].value
        cellO = "M" + str(i)
        Q11 = ws1[cellO].value
        cellO = "N" + str(i)
        Q12 = ws1[cellO].value


        Q13 = ws1["O" + str(i)].value
        Q14 = ws1["P" + str(i)].value
        Q15 = ws1["Q" + str(i)].value
        Q16 = ws1["R" + str(i)].value
        Q17 = ws1["S" + str(i)].value
        Q18 = ws1["T" + str(i)].value
        Q19 = ws1["U" + str(i)].value
        Q20 = ws1["V" + str(i)].value


        cellC = "A"+str(2)
        ans1 = ws2[cellC].value
        cellC = "B" + str(2)
        ans2 = ws2[cellC].value
        cellC = "C" + str(2)
        ans3 = ws2[cellC].value
        cellC = "D" + str(2)
        ans4 = ws2[cellC].value
        cellC = "E" + str(2)
        ans5 = ws2[cellC].value
        cellC = "F" + str(2)
        ans6 = ws2[cellC].value
        cellC = "G" + str(2)
        ans7 = ws2[cellC].value
        cellC = "H" + str(2)
        ans8 = ws2[cellC].value
        cellC = "I" + str(2)
        ans9 = ws2[cellC].value
        cellC = "J" + str(2)
        ans10 = ws2[cellC].value
        cellC = "K" + str(2)
        ans11 = ws2[cellC].value
        cellC = "L" + str(2)
        ans12 = ws2[cellC].value

        ans13 = ws2["M" + str(2)].value
        ans14 = ws2["N" + str(2)].value
        ans15 = ws2["O" + str(2)].value
        ans16 = ws2["P" + str(2)].value
        ans17 = ws2["Q" + str(2)].value
        ans18 = ws2["R" + str(2)].value
        ans19 = ws2["S" + str(2)].value
        ans20 = ws2["T" + str(2)].value


        m1 = 0
        m2 = 0
        m3 = 0

        m4 = 0
        m5 = 0

        dataT = 0

        if Q1 == ans1:
            m1 = m1 + 1
        if Q2 == ans2:
            m1 = m1 + 1
        if Q3 == ans3:
            m1 = m1 + 1
        if Q4 == ans4:
            m1 = m1 + 1

        if Q5 == ans5:
            m2 = m2 + 1
        if Q6 == ans6:
            m2 = m2 + 1
        if Q7 == ans7:
            m2 = m2 + 1
        if Q8 == ans8:
            m2 = m2 + 1

        if Q9 == ans9:
            m3 = m3 + 1
        if Q10 == ans10:
            m3 = m3 + 1
        if Q11 == ans11:
            m3 = m3 + 1
        if Q12 == ans12:
            m3 = m3 + 1


        if Q13 == ans13:
            m4 = m4 + 1
        if Q14 == ans14:
            m4 = m4 + 1
        if Q15 == ans15:
            m4 = m4 + 1
        if Q16 == ans16:
            m4 = m4 + 1

        if Q17 == ans17:
            m5 = m5 + 1
        if Q18 == ans18:
            m5 = m5 + 1
        if Q19 == ans19:
            m5 = m5 + 1
        if Q20 == ans20:
            m5 = m5 + 1

        SCORE = ws1["A"+str(i)].value

        if m1 != 0 and m2 != 0 and m3 != 0:
            if smallest(m1,m2,m3) == 1:
                SCORE = SCORE - m1
                dataT = 1
            elif smallest(m1,m2,m3) == 2:
                SCORE = SCORE - m2
                dataT = 2
            elif smallest(m1,m2,m3) == 3:
                SCORE = SCORE - m3
                dataT = 3
            else:
                print("!!!!!!!")
                print("WARNING : REPORTED -1 return error found ....")

        if m4 != 0 and m5 != 0:
            if smallest2(m4,m5) == 1:
                SCORE = SCORE - m4
                dataT = 4
            elif smallest2(m4,m5) == 2:
                SCORE = SCORE - m5
                dataT = 5
            else:
                print("!!!!!!!")
                print("WARNING : REPORTED -1 return error found ....")

        print(ws1["B" + str(i)].value+" - "+str(SCORE))


        #SAVING TO NEW EXCEL

        datasheet["A" + str(i)].value = SCORE
        datasheet["B" + str(i)].value = ws1["B" + str(i)].value
        datasheet["C" + str(i)].value = origsheet["C" + str(i)].value

        if dataT == 1:
            datasheet["X" + str(i)].value = ""
            datasheet["Y" + str(i)].value = ""
            datasheet["Z" + str(i)].value = ""
            datasheet["AA" + str(i)].value = ""
        elif dataT == 2:
            datasheet["AB" + str(i)].value = ""
            datasheet["AC" + str(i)].value = ""
            datasheet["AD" + str(i)].value = ""
            datasheet["AE" + str(i)].value = ""
        elif dataT == 3:
            datasheet["AF" + str(i)].value = ""
            datasheet["AG" + str(i)].value = ""
            datasheet["AH" + str(i)].value = ""
            datasheet["AI" + str(i)].value = ""


        if dataT == 4:
            datasheet["AJ" + str(i)].value = ""
            datasheet["AK" + str(i)].value = ""
            datasheet["AL" + str(i)].value = ""
            datasheet["AM" + str(i)].value = ""
        elif dataT == 5:
            datasheet["AN" + str(i)].value = ""
            datasheet["AO" + str(i)].value = ""
            datasheet["AP" + str(i)].value = ""
            datasheet["AQ" + str(i)].value = ""

        datasave.save('resultsXII.xlsx')



EXTRACTDATA()
